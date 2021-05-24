from openpyxl import Workbook
from PyQt5 import QtCore, QtGui, QtWidgets
from export_ui import Ui_MainWindow
from db import Connect
from datetime import datetime
from openpyxl.styles import Font,PatternFill,Alignment,Side,Border
from msgBox import box
class ExportWindow(QtWidgets.QMainWindow, Ui_MainWindow):
	def __init__(self):
		QtWidgets.QMainWindow.__init__(self)
		Ui_MainWindow.__init__(self)
		self.setupUi(self)
		self.con = Connect()
		years = []
		for year in [("All",)]+self.con.execute("SELECT datentime from ledger"):
			year = year[0].split('/')[-1]
			if year not in years:
				years.append(year)
				self.yearBox.addItem(year)
		self.exportBtn.clicked.connect(self.export)
	def display(self):
		self.show()
	def export(self):
		if self.yearBox.currentIndex()==0:
			box(2,"Failed",f"Please Select Year First","")
			return
		header = 'Al-khair cash ledger'
		headers = ['Date','Buyer','Investor','Description','Debit','Credit','Dr/Cr','Balance']
		book = Workbook()
		sheet = book.active
		sheet.name = "Al-khair Cash"
		ft = Font(size=24,bold=True)
		sheet.append([header])
		sheet['A1'].fill = PatternFill("solid",fgColor="FFFF00")
		sheet['A1'].font = ft
		sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
		sheet.merge_cells("A1:H1")
		sheet.append(headers)
		for cell in sheet['2:2']:
			cell.font = Font(size=12,bold=True)
			thin = Side(border_style="thin", color="000000")
			cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
		sheet.append([])
		sheet.row_dimensions[3].height = 2
		datum = self.con.execute("SELECT * from ledger")

		self.createTable(sheet,datum)

		sheet.append([])
		sheet.append([])
		self.addInvestors(sheet)
		filename = f"Ledger_{self.yearBox.currentText()}.xlsx"
		try:
			book.save(filename)
			box(1,"Saved",f"Ledger saved to: "+filename,"")
		except:
			box(2,"Failed",f"Close {filename} and then Try Again","")
			return
	def addInvestors(self,sheet):
		sheet.append(['Investor Name','Balance'])
		headerCell = sheet.max_row
		for cell in sheet[f'A{headerCell}:B{headerCell}'][0]:
			cell.fill = PatternFill("solid",fgColor="FFFF00")
			cell.font = Font(size=12,bold=True)
			thin = Side(border_style="thin", color="000000")
			cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

		for investor in self.con.getInvestors():
			sheet.append([investor['name'],self.con.getBalance(investor['id'])])
			self.addBorder(sheet.max_row,sheet,extra = ["A","B"])
	def addBorder(self,i,sheet, extra = None):
		cells = sheet[f'{i}:{i}'] if not extra else sheet[f'{extra[0]}{i}:{extra[1]}{i}'][0]
		for cell in cells:
			thin = Side(border_style="thin", color="000000")
			cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
	def createTable(self,sheet,datum):
		j = sheet.max_row+2
		for z,data in enumerate(datum):
			year = data[4].split('/')[-1]
			buyerId = data[1]
			investorId = data[2]
			description = data[3]
			date = data[4]
			debit = data[5]
			credit = data[6]
			buyer = ''
			arrData ={}
			arrData['date'] = date.split(' ')[0]
			arrData['buyer'] = ''
			if buyerId != -1:
				arrData["buyer"] = self.con.getBuyers(buyerId)[0]['name']
			arrData["investor"] = self.con.getInvestors(investorId)[0]['name']
			arrData['desc'] = description
			arrData['debit'] = round(debit)
			arrData['credit'] = round(credit)
			if self.yearBox.currentText() not in ['All',year]:
				continue
			if j==4:
				sheet[f'H{j-1}'].value = datum[z-1] if z>0 else 0

			i=j
			sheet.append(list(arrData.values())+[f'=IF(H{i}>0,"Dr","Cr")',f"=(H{i-1}+E{i})-F{i}"])
			self.addBorder(i,sheet)
			j+=1
import sys
if __name__ == "__main__":
	app = QtWidgets.QApplication(sys.argv)
	window = ExportWindow()
	window.show()
	sys.exit(app.exec_())