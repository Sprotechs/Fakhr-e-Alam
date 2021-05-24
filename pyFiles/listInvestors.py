from PyQt5 import QtCore, QtGui, QtWidgets
from addInvestor import AddInvestorWindow
from listInvestors_ui import Ui_MainWindow
from monthlyDetails import MonthlyDetailsWindow
from addInvestment import AddInvestmentWindow
from db import Connect
from openpyxl.styles import Font,PatternFill,Alignment,Side,Border
from openpyxl import Workbook
from msgBox import box
class ListInvestorsWindow(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self):
        QtWidgets.QMainWindow.__init__(self)
        Ui_MainWindow.__init__(self)
        self.con = Connect()
        self.setupUi(self)
        self.refresh()
        self.addBtn.clicked.connect(self.displayAddInvestorWindow)
        self.deleteBtn.clicked.connect(self.deleteInvestor)
        self.detailsBtn.clicked.connect(self.moreDetails)
        self.investmentBtn.clicked.connect(self.displayAddInvestmentWindow)
        self.saveBtn.clicked.connect(self.export)
    def export(self):
        if not self.investorsTable.selectedItems():
            box(2,"Select First","Please Select Investors","")
            return
        headers = ['Date','Description','Debit','Credit','Dr/Cr','Balance']
        book = Workbook()
        done = []

        for n,row in enumerate(self.investorsTable.selectedItems()):
            investorId = self.investorsTable.item(row.row(),0).text()
            investorName = self.investorsTable.item(row.row(),1).text()
            if investorId in done:
                continue
            done.append(investorId)
            if n==0:
                sheet = book.active
                sheet.title = investorName
            else:
                sheet=book.create_sheet(investorName)
            header = investorName
            ft = Font(size=24,bold=True)
            sheet.append([header])
            sheet['A1'].fill = PatternFill("solid",fgColor="FFFF00")
            sheet['A1'].font = ft
            sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
            sheet.merge_cells("A1:F1")
            sheet.append(headers)
            for cell in sheet['2:2']:
                cell.font = Font(size=12,bold=True)
                thin = Side(border_style="thin", color="000000")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            sheet.append([])
            sheet.row_dimensions[3].height = 2
            for year in self.con.execute("SELECT distinct(year) from monthly_details WHERE investor_id=?",(investorId,)):
                year = year[0]
                for month in self.con.execute("SELECT distinct(month) from monthly_details WHERE investor_id=? AND year=?",(investorId,year)):
                    month=month[0]
                    checkPaid = self.con.execute("SELECT amount,datentime from profit_paid WHERE investor_id=? and year=? and month=?",(investorId,year,month))
                    data = self.con.execute("SELECT profit/2 from monthly_details WHERE investor_id=? AND year=? and month=?",(investorId,year,month))
                    i = sheet.max_row+1
                    if i==3:
                        i=4
                    sheet.append([f"1/{month}/{year}",f"{month} {year} Profit",round(data[0][0]),'']+[f'=IF(F{i}>0,"Dr","Cr")',f"=(F{i-1}+C{i})-D{i}"])
                    if checkPaid:
                        sheet.append([checkPaid[0][1],f"{month} {year} Profit Paid",'',checkPaid[0][0]]+[f'=IF(F{i}>0,"Dr","Cr")',f"=(F{i-1}+C{i})-D{i}"])
                    self.addBorder(i,sheet)
            sheet.append([])
            sheet.append([])
            sheet.append(['Investment'])
            cell = sheet['A'+str(sheet.max_row)]
            cell.fill = PatternFill("solid",fgColor="FFFF00")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.merge_cells(f"A{sheet.max_row}:B{sheet.max_row}")
            cell.font = Font(size=14,bold=True)
            sheet.append(['Date','BALANCE'])
            for cell in sheet[f'A{sheet.max_row}:B{sheet.max_row}'][0]:
                cell.font = Font(size=12,bold=True)
                thin = Side(border_style="thin", color="000000")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            total = 0
            for inv in self.con.execute("SELECT join_date,amount FROM investment WHERE investor_id=?",(investorId,)):
                sheet.append(inv)
                total+=inv[1]
                for cell in sheet[f'A{sheet.max_row}:B{sheet.max_row}'][0]:
                    cell.font = Font(size=12)
                    thin = Side(border_style="thin", color="000000")
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            sheet.append(['Total',total])
            for cell in sheet[f'A{sheet.max_row}:B{sheet.max_row}'][0]:
                cell.font = Font(size=12,bold=True,color="FF0000")
                thin = Side(border_style="thin", color="000000")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        filename = f"Profit Record.xlsx"
        try:
            book.save(filename)
            box(1,"Saved",f"Profit Record saved to: "+filename,"")
        except:
            box(2,"Failed",f"Close {filename} and then Try Again","")
            return
    def addBorder(self,i,sheet, extra = None):
        cells = sheet[f'{i}:{i}'] if not extra else sheet[f'{extra[0]}{i}:{extra[1]}{i}'][0]
        for cell in cells:
            thin = Side(border_style="thin", color="000000")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    def displayAddInvestorWindow(self):
        self.addInvestorWindow = AddInvestorWindow()
        self.addInvestorWindow.display(self.refresh)
    def displayAddInvestmentWindow(self):
        self.addInvestmentWindow = AddInvestmentWindow()
        self.addInvestmentWindow.display(self.refresh)
    def display(self,button):
        self.mainRefresh = button
        self.show()
    def refresh(self):
        self.investorsTable.setRowCount(0)
        for row in self.con.getInvestors():
            self.insertRow(row)
        if self.investorsTable.rowCount()==0:
            self.investmentBtn.setEnabled(False)
            self.deleteBtn.setEnabled(False)
            self.detailsBtn.setEnabled(False)
        else:
            self.investmentBtn.setEnabled(True)
            self.deleteBtn.setEnabled(True)
            self.detailsBtn.setEnabled(True)
        try:
            self.mainRefresh()
        except:
            pass
    def insertRow(self,row):
        table = self.investorsTable
        rowPosition = table.rowCount()
        table.insertRow(rowPosition)
        moreDetails = {}
        moreDetails["total_balance"]=self.con.getBalance(int(row['id']))
        moreDetails["total_profit"]=self.con.getProfit(int(row['id']))
        moreDetails["total_investment"]=self.con.getInvestment(int(row['id']))
        table.setItem(rowPosition , 0, QtWidgets.QTableWidgetItem(row['id']))
        table.setItem(rowPosition , 1, QtWidgets.QTableWidgetItem(row['name']))
        table.setItem(rowPosition , 2, QtWidgets.QTableWidgetItem(row['address']))
        table.setItem(rowPosition , 3, QtWidgets.QTableWidgetItem(row['contact']))
        table.setItem(rowPosition , 4, QtWidgets.QTableWidgetItem(row['join_date']))
        table.setItem(rowPosition , 5, QtWidgets.QTableWidgetItem(str(round(float(moreDetails['total_balance'])))))
        table.setItem(rowPosition , 6, QtWidgets.QTableWidgetItem(str(round(float(moreDetails['total_profit'])))))
        table.setItem(rowPosition , 7, QtWidgets.QTableWidgetItem(str(round(float(moreDetails['total_investment'])))))
    def moreDetails(self):
        if not len(self.investorsTable.selectedItems()):
            return
        investorId =  self.investorsTable.item(self.investorsTable.selectedItems()[0].row(),0).text()
        self.detail = MonthlyDetailsWindow()
        self.detail.display('I',investorId)
    def finallyDelete(self,i):
        if i.text()=='Cancel':
            return
        for row in self.investorsTable.selectedItems():
            investorId = self.investorsTable.item(row.row(),0).text()
            self.con.deleteInvestor(investorId)
        self.refresh()

    def deleteInvestor(self):
        if len(self.investorsTable.selectedItems()):
            box(3,"Delete Invester",f"Are you sure you want to delete {len(self.investorsTable.selectedItems())} item(s)","",True,self.finallyDelete)
        else:
            box(2,"Select First",f"Select Invester First","")

        
import sys
if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = ListInvestorsWindow()
    window.show()
    sys.exit(app.exec_())