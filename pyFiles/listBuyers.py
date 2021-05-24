from PyQt5 import QtCore, QtGui, QtWidgets
from addBuyer import AddBuyerWindow
from db import Connect
from listBuyers_ui import Ui_MainWindow
from monthlyDetails import MonthlyDetailsWindow
from msgBox import box
from datetime import datetime
from openpyxl.styles import Font,PatternFill,Alignment,Side,Border
from openpyxl import Workbook
class ListBuyersWindow(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self):
        QtWidgets.QMainWindow.__init__(self)
        Ui_MainWindow.__init__(self)
        self.setupUi(self)
        self.buyersTable.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.con = Connect()
        self.refresh()
        self.addBtn.clicked.connect(self.displayAddBuyerWindow)
        self.deleteBtn.clicked.connect(self.deleteBuyer)
        self.detailsBtn.clicked.connect(self.moreDetails)
        self.saveBtn.clicked.connect(self.export)
    def insertRow(self,row):
        table = self.buyersTable
        rowPosition = table.rowCount()
        table.insertRow(rowPosition)
        table.setItem(rowPosition , 0, QtWidgets.QTableWidgetItem(row['id']))
        table.setItem(rowPosition , 1, QtWidgets.QTableWidgetItem(row['name']))
        table.setItem(rowPosition , 2, QtWidgets.QTableWidgetItem(row['address']))
        table.setItem(rowPosition , 3, QtWidgets.QTableWidgetItem(row['contact']))
        table.setItem(rowPosition , 4, QtWidgets.QTableWidgetItem(row['description']))
        table.setItem(rowPosition , 5, QtWidgets.QTableWidgetItem(str(round(float(row['sale_price'])))))
        table.setItem(rowPosition , 6, QtWidgets.QTableWidgetItem(str(round(float(row['remaining_price'])))))
        table.setItem(rowPosition , 7, QtWidgets.QTableWidgetItem(str(round(float(row['total_paid'])))))
    def display(self,button):
        self.mainRefresh = button
        self.show()
    def export(self):
        if not self.buyersTable.selectedItems():
            box(2,"Failed",f"First Select any Buyer","")
            return
        buyers = []
        for n,row in enumerate(self.buyersTable.selectedItems()):
            buyerId = self.buyersTable.item(row.row(),0).text()
            if buyerId in buyers:
                continue
            buyers.append(buyerId)
            buyer = self.con.getBuyers(buyerId)[0]
            buyer['investor_id'] = buyer['investor_name']
            del buyer['investor_name']
            del buyer['raff']
            months = {datetime.strptime(f"{m}",'%m').strftime('%B'):0 for m in range(1,13)}
            years = {}
            for year in self.con.execute("SELECT distinct(year) from monthly_details WHERE buyer_id=?",(buyerId,)):
                year = year[0]
                years[year] = months
                for month in self.con.execute("SELECT distinct(month) from monthly_details WHERE buyer_id=? AND year=?",(buyerId,year)):
                    month=month[0]
                    qist = self.con.execute("SELECT paid FROM monthly_details WHERE buyer_id=? AND year=? and month=?",(buyerId,year,month))
                    years[year][month] = round(qist[0][0])
            book = Workbook()
            y=0
            for year in years:
                if y==0:
                    sheet = book.active
                    sheet.title = year
                else:
                    sheet = book.create_sheet(year)
                sheet.append([r.replace('_id','_name').replace('_',' ').title() for r in buyer.keys()]+[r[:3]+"-"+year[2:] for r in years[year].keys()]+['Total'])
                d = list(buyer.values())
                total = 0
                for cell in sheet['1:1']:
                    cell.font = Font(size=12,bold=True)
                    thin = Side(border_style="thin", color="000000")
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    cell.fill = PatternFill("solid",fgColor="FFFF00")

                for month in years[year]:
                    value = years[year][month]
                    d.append(value)
                    total+=value
                sheet.append(d+[total])
                y+=1
            filename = "Buyers Detail.xlsx"
            try:
                book.save(filename)
                box(1,"Saved",f"Buyers Detail saved to: "+filename,"")
            except:
                box(2,"Failed",f"Close {filename} and then Try Again","")
                return
    def moreDetails(self):
        if not len(self.buyersTable.selectedItems()):
            box(2,"No Selection","Please Select Items First","")
            return
        buyerId =  self.buyersTable.item(self.buyersTable.selectedItems()[0].row(),0).text()
        self.detail = MonthlyDetailsWindow()
        self.detail.display('B',buyerId)

    def refresh(self):
        self.buyersTable.setRowCount(0)
        for row in self.con.getBuyers():
            self.insertRow(row)
        if self.buyersTable.rowCount()==0:
            self.deleteBtn.setEnabled(False)
            self.detailsBtn.setEnabled(False)
        else:
            self.deleteBtn.setEnabled(True)
            self.detailsBtn.setEnabled(True)
        try:
            self.mainRefresh()
        except:
            pass
    def finallyDelete(self,i):
        if i.text()=='Cancel':
            return

        for row in self.buyersTable.selectedItems():
            buyerId = self.buyersTable.item(row.row(),0).text()
            self.con.deleteBuyer(buyerId)
            self.con.deleteMonthlyDetails(buyerId)
        self.refresh()

    def deleteBuyer(self):
        if len(self.buyersTable.selectedItems()):
            box(3,"Delete Buyer",f"Are you sure you want to delete {len(self.buyersTable.selectedItems())} item(s)","",True,self.finallyDelete)
        else:
            box(2,"Select First",f"Select Buyer First","")
    def displayAddBuyerWindow(self):
        self.addBuyerWindow = AddBuyerWindow()
        self.addBuyerWindow.display(self.refresh)

import sys
if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = ListBuyersWindow()
    window.show()
    sys.exit(app.exec_())